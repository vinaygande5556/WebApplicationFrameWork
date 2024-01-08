import openpyxl
from Utilities.ReadConfigfile import get_config_data


def getRowData(file_path, sheet_name):
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.max_row


def getColumnData(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.max_column


def readData(file_path, sheet_name, row_num, col_num):
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.cell(row=row_num, column=col_num).value


def writeData(file_path, sheet_name, row_num, col_num, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    ws.cell(row=row_num, column=col_num).value = data
    wb.save(file_path)


def getInputData(sheet_name):
    path = get_config_data("Input", "path")
    rows = getRowData(path, sheet_name)
    col = getColumnData(path, sheet_name)
    mainList = []
    for i in range(2, rows + 1):
        dataList = []
        for j in range(1, col + 1):
            data = readData(path, sheet_name, i, j)
            print(data)
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    print(mainList)
    return mainList
